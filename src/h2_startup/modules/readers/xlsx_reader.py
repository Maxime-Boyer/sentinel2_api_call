from typing import Dict, List
import os

import openpyxl
from openpyxl import worksheet as Worksheet

from h2_startup.modules.readers.base_reader import BaseReader


class ExcelReader(BaseReader):
    # =============================================================================
    # user functions
    # =============================================================================
    def get_info(self, file_path: str):
        """Get information of an Excel file. Can either be a BPU or a REC document."""
        # BPU
        if "bpu" in os.path.split(file_path)[1].lower():
            return self.get_bpu_info(file_path)

        # REC
        else:
            return self.get_rec_info(file_path)

    def get_rec_info(self, file_path: str):
        """Parse REC file in Excel format and return the text of the file."""

        workbook = openpyxl.load_workbook(file_path)

        sheetnames = ["Config-Q", "Config-T", "Config-C"]

        text = ""
        balise_new_row = "\n\n"
        balise_new_page = "\n\n\n"

        for i, sheetname in enumerate(sheetnames):
            current_sheet = workbook[sheetname]
            cmpt = 1
            text += f"<Page number {i}>"

            for row in current_sheet.iter_rows():
                text += balise_new_row + str(cmpt) + ": "
                cmpt += 1
                for cell in row:
                    if cell.value is not None:
                        text += str(cell.value) + "\n"

            text += balise_new_page + f"</Page number {i}>" + balise_new_page

        return text

    def get_bpu_info(self, file_path: str) -> Dict | None:
        """Read the "BPU" file filled with informations of interest.
        Retrieve those informations in a dict.
        None is returned if no information is returned.

        Informations of interest are :
        - Catégorie de profil [code de profil]
        - Taux journaliers [€/jour]
        - Charge par profil pour chaque tranche : ferme et optionnelle(s) [jours]
        - Prix de la tranche ferme [€]
        - Nombre de jours de gratuité [jours]
        - Prix des jours de gratuité [€]
        - Prix de la tranche ferme moins les jours de gratuité [€]
        - Prix total des tranches optionnelles [€]
        - Prix total [€]

        Args:
            file_path (str): The path of the excel file.

        Returns:
            Dict: The information of interest retrieved in a dict.
            None if no information is found.
        """

        # load excel file with actual values not formulas
        sheets = openpyxl.load_workbook(file_path, data_only=True)

        # BPU's first sheet "TJM"
        sheet_tjm = sheets["TJM et charges"]
        # list of strings to search in "TJM" sheet
        strings_tjm = ["taux journaliers", "tranche ferme", "tranche optionnelle"]
        # get locations
        locations_tjm = self._search_string(sheet_tjm, strings_tjm)
        # get values from locations
        values_tjm = self._get_values_tjm(sheet_tjm, locations_tjm)

        # BPU's second sheet "BP"
        sheet_bp = sheets["Bordereau de prix"]
        # list of strings to search in "BP" sheet
        strings_bp = [
            "jours de gratuité",
            "total tranches optionnelles",
            "total tranche ferme",
        ]
        # get locations
        locations_bp = self._search_string(sheet_bp, strings_bp)
        # get values from locations
        values_bp = self._get_values_bp(sheet_bp, locations_bp)

        # return merged dict
        return {**values_tjm, **values_bp}

    # =============================================================================
    # internal functions
    # =============================================================================
    def _search_string(self, sheet: Worksheet, targets: list[str]) -> List[Dict]:
        """Search the presence of given strings inside the sheet's cells.
        Output cells location that contain one of the strings.

        Args:
            sheet: The loaded sheet of the excel file.
            targets (List[str]): The strings to search inside the sheet.

        Returns:
            Dict: Location and value of the cells containing a targeted string.
            None if no string is found.
        """
        if not all(isinstance(t, str) for t in targets):
            # test if elements of targets are strings
            return []

        locations_and_values = []

        for target in targets:
            # kill cases to facilitate quick search via string equality
            target = target.lower()
            # iterating through rows and columns
            for row in sheet.iter_rows():
                for cell in row:
                    # if the current observation is a string, we investigate it
                    if isinstance(cell.value, str):
                        # kill cases to facilitate quick search via string equality
                        current_observation = cell.value.lower()
                        # investigate and store if success
                        if target in current_observation:
                            locations_and_values.append(
                                {
                                    "value": cell.value,
                                    "row": cell.row,
                                    "column": cell.column,
                                }
                            )

        return locations_and_values

    def _get_values_bp(self, sheet: Worksheet, strings_locations: List[Dict]) -> Dict:
        # instanciate dict
        values = {}

        # loop through strings locations
        for location in strings_locations:
            # define value location with its relative position from string location
            value_row = location["row"]
            value_column = location["column"] + 2
            # get value
            value = sheet.cell(value_row, value_column).value
            # add key value to dict
            values[location["value"]] = value

        return values

    def _get_values_tjm(self, sheet: Worksheet, strings_locations: List[Dict]) -> Dict:
        # instanciate dict
        values = {}

        for location in strings_locations:
            # retrieve profile category and associated daily cost
            if "taux" in location["value"].lower():
                row = location["row"] + 1
                # get profile category value
                profil_column = location["column"] - 1
                profil_value = sheet.cell(row, profil_column).value
                values["Catégorie de profil"] = profil_value
                # get daily cost value
                tjm_column = location["column"]
                tjm_value = sheet.cell(row, tjm_column).value
                values[location["value"]] = tjm_value

            # retrieve amount of days for each "tranche"
            elif "tranche" in location["value"].lower():
                # define value location with its relative position from string location
                row = location["row"] + 2
                column = location["column"] + 1
                # get value
                value = sheet.cell(row, column).value
                # add key value to dict
                values["Nb de jours " + location["value"]] = value

        return values
